'''
John Curran

This web scraper will pull election data from Dave Leip's Election Atlas and 
organize the data into a spreadsheet or MapChart text file using Pandas. There are 
several functions for different uses.


Bibliography
Dave Leip's Atlas of U.S. Presidential Elections. http://uselectionatlas.org (17 September 2020)


required modules: requests, numpy, pandas, openpyxl, colored, BeautifulSoup
If the program does not compile, try "pip install" on these modules
'''

import requests
import numpy as np
import pandas as pd
import openpyxl
import colored
from bs4 import BeautifulSoup
from collections import defaultdict

# maps states to their corresponding FIPS value for the atlas. Alaska and Louisiana data are not avaiable.
fips = {"Alabama": 1, "Arizona": 4, "Arkansas": 5, "California": 6, "Colorado": 8, "Connecticut" : 9, "Delaware": 10, "DC": 11, "Florida": 12, "Georgia": 13, \
"Hawaii": 15, "Idaho": 16, "Illinois": 17, "Indiana": 18, "Iowa": 19, "Kansas": 20, "Kentucky": 21, "Maine": 23, "Maryland": 24, "Massachusetts": 25, "Michigan": 26, \
"Minnesota": 27, "Mississippi": 28, "Missouri": 29, "Montana": 30, "Nebraska": 31, "Nevada": 32, "New Hampshire": 33, "New Jersey": 34, "New Mexico": 35, \
"New York":36, "North Carolina":37, "North Dakota":38, "Ohio": 39, "Oklahoma": 40, "Oregon": 41, "Pennsylvania": 42, "Rhode Island": 44, "South Carolina": 45, \
"South Dakota": 46, "Tennessee": 47, "Texas": 48, "Utah": 49, "Vermont": 50, "Virginia": 51, "Washington": 53, "West Virginia": 54, "Wisconsin": 55, "Wyoming": 56}

# map of state abbreviations for use in MapCharts
abbs = {"Alabama": "AL", "Arizona": "AZ", "Arkansas": "AR", "California": "CA", "Colorado": "CO", "Connecticut" : "CT", "Delaware": "DE", "DC": "DC", \
"Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", \
"Maine": "ME", "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", \
"Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio":"OH", "Oklahoma": "OK", \
"Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT", "Vermont": "VT", \
"Virginia": "VA", "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY"}

blue = colored.fg("#5678ff")
orange = colored.fg("#ff9072")
res = colored.attr("reset")

# private method that converts a percentage string (ending with '%') into a float value
def __percent(s):
	return float(s[:-1])
	
# private method that returns whether a key exists in a dictionary
def __contains_key(dict, key):
	try:
		dict[key]
		return True
	except:
		return False
		
# private method that takes in a DataFrame row and returns the winner and the margin, in tuple form
def __winner(row):
	order = defaultdict(lambda: []) # maps the % of each candidate with the indexes associated with that %
	for v in range(1, len(row)):
		order[row[v]].append(v)
	
	keys = list(order.keys()) # list of % received by each candidate
	keys.sort(reverse=True)
	if len(order[keys[0]]) > 1: # in case of a first place tie
		return (None,0)
		
	results = [order[k][0] for k in keys] # list of column indexes for each candidate in order of vote %
	return (row.axes[0][results[0]][:-2], keys[0] - keys[1])

# private method that takes in a DataFrame row and returns the winner and the margin level, in tuple form	
def __margin(row):
	w,m = __winner(row)
	
	if m < 10:
		return (w,0)
	elif m < 20:
		return (w,1)
	elif m < 30:
		return (w,2)
	else:
		return (w,3)
	
# private method that takes in corresponding rows from each election year and computes the margin swing between elections
def __swing(row):
	swing = row["Swing"]
	if np.isnan(swing):
		return (None,None)
	
	# indexes used for the color scheme
	aswing = abs(swing)
	m = 0
	if swing == 0:
		return (2,m)
	
	if 0 < aswing < 5:
		m = 0
	elif 5 <= aswing < 10:
		m = 1
	elif 10 <= aswing < 20:
		m = 2
	elif 15 <= aswing < 30:
		m = 3
	elif 20 <= aswing < 40:
		m = 4
	elif 30 <= aswing < 50:
		m = 5
	else:
		m = 6
	
	if swing > 0: # shift towards Democrats
		return (0,m)
	else:
		return (1,m) # shift towards Republicans
	

# returns a pandas DataFrame with county level data for the election
def election_results(state, year):
	request = requests.get("https://uselectionatlas.org/RESULTS/datagraph.php?year=" + str(year) + "&fips=" + str(fips[state]))
	soup = BeautifulSoup(request.content, "html.parser")
	tables = soup.body.find("div", {"class": "info"}).find_all("table") # list of tables, each table corresponding to a county
	results = {} # will contain integer results for each county
	all_candidates = {} # running list of all candidates on the ballot in this state
	
	# add the value for every county individually
	for county in tables:
		values = county.find_all("tr") # list of candidate rows for the county		
		county_name = values[0].find("td").b.string
		
		# deal with exceptions with independent cities and counties of the same name
		for (exception,exception_state) in [("Baltimore","Maryland"), ("Fairfax","Virginia"), ("Richmond", "Virginia"), \
		("Bedford", "Virginia"), ("Franklin", "Virginia"), ("Roanoke","Virginia"), ("St. Louis", "Missouri")]:
			if county_name == exception and state == exception_state:
				if not __contains_key(results, exception + " County"):
					county_name = exception + " County"
				else:
					county_name = exception + " City"
				break
					
		if county_name == "District of Columbia": # Washington DC edge case
			county_name = "Washington"
		# small changes in county names
		elif county_name == "Dewitt" and state == "Texas":
			county_name = "DeWitt"
		elif county_name == "Desoto" and state == "Florida":
			county_name = "DeSoto"
		elif county_name == "Dade" and state == "Florida":
			county_name = "Miami-Dade"
		elif county_name == "Ormsby" and state == "Nevada":
			county_name = "Carson City"
		elif county_name == "Shannon" and state == "South Dakota":
			county_name = "Oglala Lakota"
		
		candidate_results = defaultdict(lambda: 0.0) # results for each candidate in the county
		for candidate in values:
			name = candidate.find("td", {"class":"cnd"})
			if name is None:
				name = candidate.find("td").string
			else:
				name = name.string
				
			candidate_results[name] = __percent(candidate.find("td", {"class":"per"}).string)
			all_candidates[name] = None
			
		results[county_name] = candidate_results
		
	'''
	Current data format for results
	{
	county_name:
		{Candidate: percentage, Candidate: percentage, etc},
	county_name:
		{Candidate: percentage, Candidate: percentage, etc},
	county_name:
		{Candidate: percentage, Candidate: percentage, etc},
	etc
	}
	'''
		
	# return results in the form of a DataFrame
	df = {"County": results.keys()}
	for candidate in all_candidates.keys(): # for every candidate
		df[candidate + " %"] = [results[county][candidate] for county in results.keys()]
		
	return pd.DataFrame(df)
	

# # returns a pandas DataFrame with county level data for swings between election years in a state
def election_swings(state, year1, year2):
	results1 = election_results(state, year1)
	results2 = election_results(state, year2)
	results = results1.merge(results2, on="County", how="outer")
	candidates = results.columns
	
	# calculate swing
	swing = []
	for i in range(len(results)):
		swing += [(results[candidates[len(results1.columns)]][i] - results[candidates[len(results1.columns) + 1]][i]) - (results[candidates[1]][i] - results[candidates[2]][i])]
	results["Swing"] = swing
	
	return results


# functions used with election_results


# creates an excel sheet displaying the election data
def create_simple_excel(state, year, file_name=None):
	if file_name is None:
		file_name = state + "_" + str(year) + "_Results.xlsx"
		
	results = election_results(state, year)
	results.to_excel(file_name, index=False)
	print(blue + "Success! " + file_name + " has been created." + res)


# creates an excel sheet displaying the election data for both elections and the swing between elections.
# a negative swing is a swing towards Republicans, while a positive swing is a swing towards Democrats
def create_swing_excel(state, year1, year2, file_name=None):
	if file_name is None:
		file_name = state + "_" + str(year1) + "_to_" + str(year2) + "_Swing.xlsx"
	
	results1 = election_results(state, year1)
	results = election_swings(state, year1, year2)
	candidates = results.columns
	
	results.to_excel(file_name, index=False, startrow=1)
	
	wb = openpyxl.load_workbook(file_name)
	ws = wb["Sheet1"]
	cell1 = ws.cell(1,2)
	cell2 = ws.cell(1, len(results1.columns) + 1)
	cell1.value = str(year1)
	cell2.value = str(year2)
	ws.insert_cols(len(results.columns))
	wb.save(file_name)
	
	print(blue + "Success! " + file_name + " has been created." + res)


'''
MapChart data format to be replicated

{
"groups":
	{
	hex color:
		{"div":"box0","label":color title,"paths":[name__stateabrreviation,name__state,...]},
	hex color:
		{"div":"box1","label":color title,"paths":[name__state,name__state,...]},
	etc
	},
"title":title,"hidden":[],"background":"#ffffff","borders":"#000000","legendFont":"Century Gothic","legendFontColor":"#000000","legendBgColor":"rgba(0, 0, 0, 0)"
}
'''


# takes in a list of states (or a single state) and a year and produces a mapchart document for the election results
# of that year in the given states
def create_simple_mapchart(states, year, title=None, file_name=None, colors=[]):
	if title is None:
		title = str(year) + " Election"			
	if file_name is None:
		file_name = str(year) + "_Results_MapChart.txt"
	if not type(states) is list:
		states = [states]
			
	# set default colors if they were not put in by user
	clist = ["#0064ff", "#ff4545", "#19d500", "#ffd100"]
	if len(colors) < 4:
		for i in range(len(colors), 4):
			colors.append(clist[i])
		
	results = [election_results(state, year) for state in states]
	
	all_candidates = {} # running list of all candidates in the country
	for state in results:
		for cand in state.columns[1:]:
			all_candidates[cand[:-2]] = None
	all_candidates = list(all_candidates.keys())
	
	chart = {"groups": {},"title":title,"hidden":[],"background":"#ffffff", "borders":"#000000","legendFont":"Century Gothic",\
	"legendFontColor":"#000000","legendBgColor":"rgba (0, 0, 0, 0)"}
	
	color_map = {} # color paths
	for i in range(len(all_candidates)): # place all candidates in the legend
		color_map[all_candidates[i]] = colors[i]
		chart["groups"][colors[i]] = {"div":"#box" + str(i),"label":all_candidates[i],"paths":[]}

	# insert the data into the chart
	warnings = "" # list of counties that are tied and thus will not be colored
	for s in range(len(states)): # iterate through all states
		for r in range(len(results[s])): # iterate through all rows
			w,_ = __winner(results[s].loc[r])
			if not w is None:
				chart["groups"][color_map[w]]["paths"].append(results[s]["County"][r].replace(" ", "_").replace("'","_").replace(".","_").replace("-","_") + "__" + abbs[states[s]])
			else:
				warnings += "Warning: " + results[s]["County"][r] + ", " + abbs[states[s]] + " was a tie.\n"
			
	# clear candidates that did not win any counties
	keys = []
	for color in chart["groups"].keys():
		if chart["groups"][color]["paths"] == []:
			keys.append(color)
			
	for color in keys:
		chart["groups"].pop(color)
			
	# create mapchart text file
	with open(file_name, "w") as f:
		ftext = str(chart).replace("'", '"') # mapchart text
		# deal with exceptions and inconsistencies in how MapChart interprets text documents
		ftext = ftext.replace("Lac_Qui", "Lac_qui").replace("_County_", "_Co__").replace("Baltimore_Co___MD", "Baltimore_County__MD")
		ftext = ftext.replace("Baltimore__MD", "Baltimore_City__MD").replace("St__Mary_s__MD", "St_Mary_s__MD")
		ftext = ftext.replace("Ste__Genevieve", "Sainte_Genevieve").replace("LaSalle", "La_Salle")
		ftext = ftext.replace("Fairfax_City", "Fairfax").replace("Richmond_City", "Richmond").replace("Bedford_City", "Bedford")
		ftext = ftext.replace("Franklin_City", "Franklin").replace("Roanoke_City", "Roanoke").replace("St__Louis_City", "St__Louis")
		f.write(ftext)
		f.close()
		
	print(blue + "Success! " + file_name + " has been created." + res)
	print(orange + warnings + res, end="")
	
	
def create_swing_mapchart(states, year1, year2, title=None, file_name=None, colors=[]):
	if title is None:
		title = str(year1) + " > " + str(year2) + " Election Swings"			
	if file_name is None:
		file_name = str(year1) + "_" + str(year2) + "_Swing_MapChart.txt"
	if not type(states) is list:
		states = [states]
			
	# set default colors if they were not put in by user
	clist = [["#58faff", "#74b7ff", "#3d92ff", "#1a55ff", "#0034c7", "#002489", "#001655"], \
	["#fccde5", "#ff9fb5", "#ff8c8c", "#ff3f3f", "#c70000", "#7e0000", "#420000"], ["#ffe436"]]
	if len(colors) < 3:
		for i in range(len(colors), 3):
			colors.append(clist[i])			
	for i in range(2):
		if len(colors[i]) < 7:
			for j in range(len(colors[i]), 7):
				colors[i].append(clist[i][j])

	results = [election_swings(state, year1, year2) for state in states]
	
	chart = {"groups": {},"title":title,"hidden":[],"background":"#ffffff", "borders":"#000000","legendFont":"Century Gothic",\
	"legendFontColor":"#000000","legendBgColor":"rgba (0, 0, 0, 0)"}
	
	all_swings = {"D > 50": colors[0][6], "D > 40": colors[0][5], "D > 30": colors[0][4], "D > 20": colors[0][3], "D > 10": colors[0][2], \
	"D > 5": colors[0][1], "D > 0": colors[0][0], "No Shift": colors[2][0], "R > 0": colors[1][0], "R > 5": colors[1][1],"R > 10": colors[1][2], \
	"R > 20": colors[1][3],"R > 30": colors[1][4],"R > 40": colors[1][5],"R > 50": colors[1][6]}
	
	box = 0
	for k in all_swings.keys(): # place all swings in the legend
		chart["groups"][all_swings[k]] = {"div":"#box" + str(box),"label":k,"paths":[]}
		box += 1

	# insert the data into the chart
	warnings = "" # for missing data
	for s in range(len(states)): # iterate through all states
		for r in range(len(results[s])): # iterate through all rows
			w,m = __swing(results[s].loc[r])
			if not w is None:
				chart["groups"][colors[w][m]]["paths"].append(results[s]["County"][r].replace(" ", "_").replace("'","_").replace(".","_").replace("-","_") + "__" + abbs[states[s]])
			else:
				warnings += "Warning: No swing data collected for " + results[s]["County"][r] + ", " + states[s] + ".\n"
			
	# clear swings that did not occur
	keys = []
	for color in chart["groups"].keys():
		if chart["groups"][color]["paths"] == []:
			keys.append(color)
			
	for color in keys:
		chart["groups"].pop(color)
			
	# create mapchart text file
	with open(file_name, "w") as f:
		ftext = str(chart).replace("'", '"') # mapchart text
		# deal with exceptions and inconsistencies in how MapChart interprets text documents
		ftext = ftext.replace("Lac_Qui", "Lac_qui").replace("_County_", "_Co__").replace("Baltimore_Co___MD", "Baltimore_County__MD")
		ftext = ftext.replace("Baltimore__MD", "Baltimore_City__MD").replace("St__Mary_s__MD", "St_Mary_s__MD")
		ftext = ftext.replace("Ste__Genevieve", "Sainte_Genevieve").replace("LaSalle", "La_Salle")
		ftext = ftext.replace("Fairfax_City", "Fairfax").replace("Richmond_City", "Richmond").replace("Bedford_City", "Bedford")
		ftext = ftext.replace("Franklin_City", "Franklin").replace("Roanoke_City", "Roanoke").replace("St__Louis_City", "St__Louis")
		f.write(ftext)
		f.close()
		
	print(blue + "Success! " + file_name + " has been created." + res)
	print(orange + warnings + res, end="")
	
	
def create_margin_mapchart(states, year, title=None, file_name=None, colors=[]):
	if title is None:
		title = str(year) + " Election"			
	if file_name is None:
		file_name = str(year) + "_Results_MapChart.txt"
	if not type(states) is list:
		states = [states]
			
	# set default colors if they were not put in by user
	clist = [["#58faff", "#5f9eff", "#0037ff", "#0023a0"], ["#fccde5", "#ff9090", "#ff2b2b", "#980000"], \
	["#afffa5", "#4dff36", "#16b800", "#0f8200"], ["#ffffae", "#ffff3d", "#ffaf00", "#d16100"]]
	if len(colors) < 4:
		for i in range(len(colors), 4):
			colors.append(clist[i])			
	for i in range(4):
		if len(colors[i]) < 4:
			for j in range(len(colors[i]), 4):
				colors[i].append(clist[i][j])
		
	results = [election_results(state, year) for state in states]
	
	all_candidates = {} # running list of all candidates in the country
	for state in results:
		for cand in state.columns[1:]:
			all_candidates[cand[:-2]] = None
	all_candidates = list(all_candidates.keys())
	
	chart = {"groups": {},"title":title,"hidden":[],"background":"#ffffff", "borders":"#000000","legendFont":"Century Gothic",\
	"legendFontColor":"#000000","legendBgColor":"rgba (0, 0, 0, 0)"}
	
	color_map = {} # color paths
	j = 0
	for i in range(len(all_candidates)): # place all candidates in the legend
		color_map[all_candidates[i]] = colors[i]
		chart["groups"][colors[i][3]] = {"div":"#box" + str(j),"label":all_candidates[i] + " > 30%","paths":[]}
		chart["groups"][colors[i][2]] = {"div":"#box" + str(j+1),"label":all_candidates[i] + " > 20%","paths":[]}
		chart["groups"][colors[i][1]] = {"div":"#box" + str(j+2),"label":all_candidates[i] + " > 10%","paths":[]}
		chart["groups"][colors[i][0]] = {"div":"#box" + str(j+3),"label":all_candidates[i] + " > 0%","paths":[]}
		j += 4

	# insert the data into the chart
	warnings = "" # list of counties that are tied and thus will not be colored
	for s in range(len(states)): # iterate through all states
		for r in range(len(results[s])): # iterate through all rows
			w,m = __margin(results[s].loc[r])
			if not w is None:
				chart["groups"][color_map[w][m]]["paths"].append(results[s]["County"][r].replace(" ", "_").replace("'","_").replace(".","_").replace("-","_") + "__" + abbs[states[s]])
			else:
				warnings += "Warning: " + results[s]["County"][r] + ", " + abbs[states[s]] + " was a tie.\n"
			
	# clear candidates that did not win any counties
	keys = []
	for color in chart["groups"].keys():
		if chart["groups"][color]["paths"] == []:
			keys.append(color)
			
	for color in keys:
		chart["groups"].pop(color)
			
	# create mapchart text file
	with open(file_name, "w") as f:
		ftext = str(chart).replace("'", '"') # mapchart text
		# deal with exceptions and inconsistencies in how MapChart interprets text documents
		ftext = ftext.replace("Lac_Qui", "Lac_qui").replace("_County_", "_Co__").replace("Baltimore_Co___MD", "Baltimore_County__MD")
		ftext = ftext.replace("Baltimore__MD", "Baltimore_City__MD").replace("St__Mary_s__MD", "St_Mary_s__MD")
		ftext = ftext.replace("Ste__Genevieve", "Sainte_Genevieve").replace("LaSalle", "La_Salle")
		ftext = ftext.replace("Fairfax_City", "Fairfax").replace("Richmond_City", "Richmond").replace("Bedford_City", "Bedford")
		ftext = ftext.replace("Franklin_City", "Franklin").replace("Roanoke_City", "Roanoke").replace("St__Louis_City", "St__Louis")
		f.write(ftext)
		f.close()
		
	print(blue + "Success! " + file_name + " has been created." + res)
	print(orange + warnings + res, end="")

